"""
Product Import/Export Functions.
Handles CSV and Excel import/export for products.
"""
import csv
import io
import json
import uuid
from datetime import datetime
from typing import TYPE_CHECKING

from fastapi import UploadFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session

from app.models.product import Product, ProductCategory

if TYPE_CHECKING:
    pass


class ImportResult(BaseModel):
    """Result of import operation."""
    success: bool
    total_rows: int
    imported: int
    updated: int
    skipped: int
    errors: list[str]


# CSV Export Headers
# Required: name, price, category_id
# Optional: description, stock, is_available, ingredients, calories, image_url, hero_image, bottle_image, thumbnail_image
EXPORT_HEADERS = [
    "name", "description", "price", "category_id", "stock",
    "is_available", "ingredients", "calories", "image_url",
    "hero_image", "bottle_image", "thumbnail_image"
]

# Minimum required headers for import (images are optional)
REQUIRED_IMPORT_HEADERS = ["name", "price", "category_id"]

# Column aliases for flexible imports (maps various header names to standard names)
# This allows case-insensitive and alternative header names
COLUMN_ALIASES = {
    # name variations
    "name": "name",
    "nama": "name",
    "nama produk": "name",
    "product name": "name",
    "product_name": "name",
    "productname": "name",
    # description variations
    "description": "description",
    "deskripsi": "description",
    "desc": "description",
    "keterangan": "description",
    # price variations
    "price": "price",
    "harga": "price",
    "base_price": "price",
    "baseprice": "price",
    # category_id variations
    "category_id": "category_id",
    "categoryid": "category_id",
    "category": "category_id",
    "kategori": "category_id",
    "kategori_id": "category_id",
    # stock variations
    "stock": "stock",
    "stok": "stock",
    "stock_quantity": "stock",
    "stockquantity": "stock",
    "quantity": "stock",
    "qty": "stock",
    # is_available variations
    "is_available": "is_available",
    "isavailable": "is_available",
    "available": "is_available",
    "tersedia": "is_available",
    "aktif": "is_available",
    "active": "is_available",
    # ingredients variations
    "ingredients": "ingredients",
    "bahan": "ingredients",
    "ingredient": "ingredients",
    "bahan-bahan": "ingredients",
    # calories variations
    "calories": "calories",
    "kalori": "calories",
    "calorie": "calories",
    # image_url variations
    "image_url": "image_url",
    "imageurl": "image_url",
    "image": "image_url",
    "gambar": "image_url",
    # hero_image variations
    "hero_image": "hero_image",
    "heroimage": "hero_image",
    "hero": "hero_image",
    # bottle_image variations
    "bottle_image": "bottle_image",
    "bottleimage": "bottle_image",
    "bottle": "bottle_image",
    # thumbnail_image variations
    "thumbnail_image": "thumbnail_image",
    "thumbnailimage": "thumbnail_image",
    "thumbnail": "thumbnail_image",
    "thumb": "thumbnail_image",
}


def _normalize_header(header: str) -> str:
    """Normalize header to standard name using aliases."""
    if not header:
        return ""
    # Convert to lowercase and strip whitespace
    normalized = header.strip().lower()
    # Remove extra whitespace
    normalized = " ".join(normalized.split())
    # Replace underscores with nothing for matching
    normalized_no_underscore = normalized.replace("_", "")
    
    # Check exact match first
    if normalized in COLUMN_ALIASES:
        return COLUMN_ALIASES[normalized]
    
    # Check without underscores
    if normalized_no_underscore in COLUMN_ALIASES:
        return COLUMN_ALIASES[normalized_no_underscore]
    
    # Return original lowercase if no match found
    return normalized


def _normalize_row_headers(row: dict) -> dict:
    """Normalize all headers in a row dictionary."""
    normalized = {}
    for key, value in row.items():
        if key is not None:
            norm_key = _normalize_header(str(key))
            # Only use the first occurrence if there are duplicates
            if norm_key not in normalized:
                normalized[norm_key] = value
    return normalized


def _get_product_ingredients_str(product: Product) -> str:
    """Get product ingredients as semicolon-separated string."""
    if not product.ingredients:
        return ""
    try:
        return ";".join(json.loads(product.ingredients))
    except json.JSONDecodeError:
        return ""


def _parse_ingredients(ingredients_str: str) -> str | None:
    """Parse ingredients string to JSON."""
    if not ingredients_str:
        return None
    return json.dumps([i.strip() for i in ingredients_str.split(";") if i.strip()])


def _product_to_csv_row(product: Product) -> list:
    """Convert product to CSV row."""
    return [
        product.name,
        product.description or "",
        product.base_price,
        product.category_id,
        product.stock_quantity,
        "true" if product.is_available else "false",
        _get_product_ingredients_str(product),
        product.calories or "",
        product.image_url or "",
        product.hero_image or "",
        product.bottle_image or "",
        product.thumbnail_image or "",
    ]


# ============================================================
# CSV Export
# ============================================================

def export_products_to_csv(products: list[Product]) -> StreamingResponse:
    """Export products to CSV format."""
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Write header
    writer.writerow(EXPORT_HEADERS)
    
    # Write data
    for product in products:
        writer.writerow(_product_to_csv_row(product))
    
    output.seek(0)
    
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={
            "Content-Disposition": f"attachment; filename=products_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        }
    )


def generate_csv_template() -> StreamingResponse:
    """Generate CSV template for product import.
    
    FLEXIBLE IMPORT SYSTEM:
    =======================
    ✅ Case-insensitive headers: "Name", "NAME", "name" all work
    ✅ Indonesian headers supported: "Nama", "Harga", "Kategori", "Stok", etc.
    ✅ Partial imports: You don't need all columns, only the required ones
    
    REQUIRED COLUMNS (at minimum you need these):
    - name (or: Nama, Product Name, Nama Produk)
    - price (or: Harga, Base Price)
    - category_id (or: Category, Kategori)
    
    OPTIONAL COLUMNS (will use defaults if missing):
    - description → defaults to "Delicious {name} juice"
    - stock → defaults to 100
    - is_available → defaults to true
    - ingredients → defaults to empty
    - calories → defaults to null
    - image_url → defaults to "bg-green-500"
    - hero_image, bottle_image, thumbnail_image → defaults to null
    
    CATEGORY IDs AVAILABLE:
    - fruit-juices: Jus Buah
    - vegetable-juices: Jus Sayur  
    - smoothies: Smoothies
    - health-shots: Health Shots
    - specialty: Spesial
    """
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Write instruction comment rows
    writer.writerow(["# TEMPLATE IMPORT PRODUK JUICEQU"])
    writer.writerow(["# ================================"])
    writer.writerow(["# Kolom WAJIB: name, price, category_id"])
    writer.writerow(["# Kolom lain OPSIONAL (bisa dikosongkan atau dihapus)"])
    writer.writerow(["# Header case-insensitive: 'Name', 'NAME', 'name' semua valid"])
    writer.writerow(["# Hapus baris komentar ini (#) sebelum import!"])
    writer.writerow([])
    
    # Write header
    writer.writerow(EXPORT_HEADERS)
    
    # Write example rows showing required vs optional fields
    # Example 1: Full data with all columns
    writer.writerow([
        "Jus Jeruk Segar",
        "Jus jeruk segar yang menyegarkan dan kaya vitamin C",
        15000,
        "fruit-juices",
        100,
        "true",
        "Jeruk;Air;Gula",
        120,
        "bg-orange-500",
        "",
        "",
        ""
    ])
    
    # Example 2: Minimal required data only
    writer.writerow([
        "Jus Apel",
        "",  # description - kosong, akan auto-generate
        12000,
        "fruit-juices",
        "",  # stock - kosong, default 100
        "",  # is_available - kosong, default true
        "",  # ingredients - opsional
        "",  # calories - opsional
        "",  # image_url - kosong, default bg-green-500
        "",  # hero_image - opsional
        "",  # bottle_image - opsional
        "",  # thumbnail_image - opsional
    ])
    
    # Example 3: Different category
    writer.writerow([
        "Green Smoothie",
        "Smoothie sehat dengan bayam dan pisang",
        18000,
        "smoothies",
        50,
        "true",
        "Bayam;Pisang;Yogurt;Madu",
        95,
        "bg-green-600",
        "",
        "",
        ""
    ])
    
    output.seek(0)
    
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={
            "Content-Disposition": "attachment; filename=products_import_template.csv"
        }
    )


# ============================================================
# Excel Export
# ============================================================

def export_products_to_excel(products: list[Product]) -> StreamingResponse:
    """Export products to Excel format (XLSX)."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        from app.core.exceptions import BadRequestException
        raise BadRequestException("Excel export requires openpyxl library. Please install it.")
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Products"
    
    # Style for headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4A5568", end_color="4A5568", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write headers with styling
    for col, header in enumerate(EXPORT_HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Write data
    for row_idx, product in enumerate(products, 2):
        data = _product_to_csv_row(product)
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
    
    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except (TypeError, AttributeError):
                # Skip cells with invalid values
                continue
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width    
    # Save to memory
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=products_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        }
    )


# ============================================================
# CSV Import
# ============================================================

def _validate_and_process_row(
    row: dict,
    row_num: int,
    db: Session,
    result: dict
) -> Product | None:
    """Validate and process a single import row."""
    name = str(row.get("name", "") or "").strip()
    if not name:
        result["errors"].append(f"Row {row_num}: Missing product name")
        result["skipped"] += 1
        return None
    
    description = str(row.get("description", "") or "").strip()
    if not description or len(description) < 10:
        description = f"Delicious {name} juice"
    
    try:
        price = float(row.get("price", 0) or 0)
        if price <= 0:
            result["errors"].append(f"Row {row_num}: Invalid price for '{name}'")
            result["skipped"] += 1
            return None
    except (ValueError, TypeError):
        result["errors"].append(f"Row {row_num}: Invalid price format for '{name}'")
        result["skipped"] += 1
        return None
    
    category_id = str(row.get("category_id", "") or "").strip()
    if not category_id:
        result["errors"].append(f"Row {row_num}: Missing category_id for '{name}'")
        result["skipped"] += 1
        return None
    
    # Validate category exists
    category = db.query(ProductCategory).filter(ProductCategory.id == category_id).first()
    if not category:
        result["errors"].append(f"Row {row_num}: Invalid category_id '{category_id}' for '{name}'")
        result["skipped"] += 1
        return None
    
    try:
        stock = int(row.get("stock", 100) or 100)
    except (ValueError, TypeError):
        stock = 100
    
    is_available_str = str(row.get("is_available", "true") or "true").strip().lower()
    is_available = is_available_str in ("true", "1", "yes")
    
    ingredients = _parse_ingredients(str(row.get("ingredients", "") or ""))
    
    try:
        calories = int(row.get("calories", 0) or 0) if row.get("calories") else None
    except (ValueError, TypeError):
        calories = None
    
    # Check if product exists
    existing = db.query(Product).filter(Product.name.ilike(name)).first()
    
    if existing:
        # Update existing product
        existing.description = description
        existing.base_price = price
        existing.category_id = category_id
        existing.stock_quantity = stock
        existing.is_available = is_available
        existing.is_deleted = False  # Restore if was deleted
        if ingredients:
            existing.ingredients = ingredients
        if calories:
            existing.calories = calories
        # Only update image fields if they have non-empty values
        # This allows importing data without images (images can be added manually later)
        if row.get("image_url") and str(row.get("image_url")).strip():
            existing.image_url = str(row.get("image_url")).strip()
        if row.get("hero_image") and str(row.get("hero_image")).strip():
            existing.hero_image = str(row.get("hero_image")).strip()
        if row.get("bottle_image") and str(row.get("bottle_image")).strip():
            existing.bottle_image = str(row.get("bottle_image")).strip()
        if row.get("thumbnail_image") and str(row.get("thumbnail_image")).strip():
            existing.thumbnail_image = str(row.get("thumbnail_image")).strip()
        result["updated"] += 1
        return existing
    else:
        # Create new product
        # Note: All image fields are optional
        # - image_url defaults to "bg-green-500" (placeholder color) if not provided
        # - hero_image, bottle_image, thumbnail_image default to None if not provided
        # Images can be uploaded later through the admin panel
        new_product = Product(
            id=str(uuid.uuid4()),
            name=name,
            description=description,
            base_price=price,
            category_id=category_id,
            stock_quantity=stock,
            is_available=is_available,
            ingredients=ingredients,
            calories=calories,
            image_url=str(row.get("image_url")).strip() if row.get("image_url") and str(row.get("image_url")).strip() else "bg-green-500",
            hero_image=str(row.get("hero_image")).strip() if row.get("hero_image") and str(row.get("hero_image")).strip() else None,
            bottle_image=str(row.get("bottle_image")).strip() if row.get("bottle_image") and str(row.get("bottle_image")).strip() else None,
            thumbnail_image=str(row.get("thumbnail_image")).strip() if row.get("thumbnail_image") and str(row.get("thumbnail_image")).strip() else None,
        )
        db.add(new_product)
        result["imported"] += 1
        return new_product


# Maximum file size for imports (10 MB)
MAX_FILE_SIZE = 10 * 1024 * 1024


async def import_products_from_csv(file: UploadFile, db: Session) -> dict:
    """Import products from CSV file.
    
    Features:
    - Case-insensitive column headers (Name, NAME, name all work)
    - Supports Indonesian and English column names
    - Only requires: name, price, category_id
    - All other columns are optional with smart defaults
    - Skips comment lines (starting with #) and empty rows
    """
    from app.core.exceptions import BadRequestException
    
    if not file.filename.endswith('.csv'):
        raise BadRequestException("File must be a CSV file")
    
    content = await file.read()
    
    # Validate file size
    if len(content) > MAX_FILE_SIZE:
        raise BadRequestException(
            f"File size exceeds maximum allowed size of {MAX_FILE_SIZE // (1024 * 1024)} MB"
        )
    
    try:
        text = content.decode('utf-8')
    except UnicodeDecodeError:
        try:
            text = content.decode('latin-1')
        except:
            raise BadRequestException("Unable to decode file. Please use UTF-8 encoding.")
    
    # Filter out comment lines (starting with #) and empty lines before parsing
    lines = text.strip().split('\n')
    filtered_lines = []
    for line in lines:
        stripped = line.strip()
        # Skip empty lines and comment lines
        if not stripped or stripped.startswith('#'):
            continue
        filtered_lines.append(line)
    
    if len(filtered_lines) < 2:
        raise BadRequestException("CSV file must have at least a header row and one data row")
    
    filtered_text = '\n'.join(filtered_lines)
    reader = csv.DictReader(io.StringIO(filtered_text))
    
    result = {
        "total_rows": 0,
        "imported": 0,
        "updated": 0,
        "skipped": 0,
        "errors": []
    }
    
    for row_num, row in enumerate(reader, 2):
        # Skip empty rows (all values are empty)
        if not any(str(v).strip() for v in row.values() if v is not None):
            continue
            
        result["total_rows"] += 1
        
        try:
            # Normalize all headers to standard lowercase names
            normalized_row = _normalize_row_headers(row)
            _validate_and_process_row(normalized_row, row_num, db, result)
        except Exception as e:
            result["errors"].append(f"Row {row_num}: Unexpected error - {str(e)}")
            result["skipped"] += 1
    
    db.commit()
    
    return {
        "success": True,
        "total_rows": result["total_rows"],
        "imported": result["imported"],
        "updated": result["updated"],
        "skipped": result["skipped"],
        "errors": result["errors"][:20]  # Limit errors to first 20
    }


# ============================================================
# Excel Import
# ============================================================

async def import_products_from_excel(file: UploadFile, db: Session) -> dict:
    """Import products from Excel file.
    
    Features:
    - Case-insensitive column headers (Name, NAME, name all work)
    - Supports Indonesian and English column names
    - Only requires: name, price, category_id
    - All other columns are optional with smart defaults
    """
    from app.core.exceptions import BadRequestException
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise BadRequestException("File must be an Excel file (.xlsx or .xls)")
    
    try:
        import openpyxl
    except ImportError:
        raise BadRequestException("Excel import requires openpyxl library. Please install it.")
    
    content = await file.read()
    
    # Validate file size
    if len(content) > MAX_FILE_SIZE:
        raise BadRequestException(
            f"File size exceeds maximum allowed size of {MAX_FILE_SIZE // (1024 * 1024)} MB"
        )
    
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
    except Exception as e:
        raise BadRequestException(f"Unable to read Excel file: {str(e)}")
    
    result = {
        "total_rows": 0,
        "imported": 0,
        "updated": 0,
        "skipped": 0,
        "errors": []
    }
    
    # Get headers from first row and normalize them
    original_headers = [cell.value for cell in ws[1] if cell.value]
    if not original_headers:
        raise BadRequestException("Excel file has no headers in the first row")
    
    # Normalize headers for flexible matching
    headers = [_normalize_header(str(h)) if h else "" for h in original_headers]
    
    # Process rows
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        # Skip empty rows
        if not any(row):
            continue
        
        result["total_rows"] += 1
        
        # Create dict from row with normalized headers
        row_dict = {}
        for idx, header in enumerate(headers):
            if idx < len(row) and header:
                row_dict[header] = row[idx]
        
        try:
            _validate_and_process_row(row_dict, row_num, db, result)
        except Exception as e:
            result["errors"].append(f"Row {row_num}: Unexpected error - {str(e)}")
            result["skipped"] += 1
    
    db.commit()
    
    return {
        "success": True,
        "total_rows": result["total_rows"],
        "imported": result["imported"],
        "updated": result["updated"],
        "skipped": result["skipped"],
        "errors": result["errors"][:20]
    }
